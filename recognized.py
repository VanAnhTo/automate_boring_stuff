from ocr import ocr_space_file
import json
import openpyxl
import os
import shutil
import datetime

path = "input"
dirs = os.listdir( path )

#List of languages will be tested
sheet_languages = {'eng': 'English', 'jpn': 'Japanese', 'rus': 'Russian', 'gre': 'Greek'}
        
wb = openpyxl.load_workbook('LanguageTest.xlsx')

backup_folder = "backup"
os.makedirs(backup_folder, exist_ok=True)

for folder in dirs:
    
    folder_name = folder    
    print(folder)
    files = os.listdir( path + "/" + folder )
    sheet = wb.get_sheet_by_name(sheet_languages[folder])

    i = 2
    
    for file in files:
        print(file)
        file_name = path + "/" + folder + "/" + file
        test_file = ocr_space_file(file_name, language = folder)
        json_obj = json.loads(test_file)
        #print(json_obj)

        actual_result = ""        
        actual_result = json_obj["ParsedResults"][0]["ParsedText"].strip().replace("\r\n", "")

        print ('Actual: ' + actual_result)
    
        expected_result = sheet.cell(row=i, column=1).value
    
    
        sheet['B'+ str(i)] = actual_result

        #print('Expected: ' + expected_result)


  
        if  expected_result == actual_result:
           sheet.cell(row=i, column=3).value = 'Pass'
        else:
           sheet.cell(row=i, column=3).value = 'Fail'

        print (sheet.cell(row=i, column=3).value)

        #get time now and change file name = run time
           
        now = datetime.datetime.now()
        new_file_name = now.strftime(folder+"_"+"%Y%m%d_%H%M%S"+ ".png")

        #add hyperlink to excel and move the file to backup folder
        print(new_file_name)
        sheet.cell(row=i, column=4).hyperlink = backup_folder + "/" + new_file_name

        shutil.move(file_name, backup_folder + "/" + new_file_name)

        i= i+1
    
wb.save('LanguageTest.xlsx')
    
