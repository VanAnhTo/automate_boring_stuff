from ocr import ocr_space_file
import json
import openpyxl
import os
import shutil
import datetime

now = datetime.datetime.now()


file_name = '2.jpg'
test_file = ocr_space_file(filename=file_name, language='jpn')

json_obj = json.loads(test_file)

actual_result = ""
for key in json_obj["ParsedResults"][0]["TextOverlay"]["Lines"]:
    actual_result += key["Words"][0]["WordText"]
    
actual_result = json_obj["ParsedResults"][0]["ParsedText"]
print ('Actual: ' + actual_result)


wb = openpyxl.load_workbook('example.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')

expected_result = sheet.cell(row=2, column=1).value
sheet['B2'] = actual_result

#print (sheet.cell(row=2, column=2).value)

if  expected_result == actual_result:
   sheet.cell(row=2, column=3).value = 'Pass'
else:
   sheet.cell(row=2, column=3).value = 'Fail'
wb.save('example.xlsx')

folder_name = "backup"
os.makedirs(folder_name, exist_ok=True)
new_file_name = now.strftime("%Y%m%d%H%M%S.png")
shutil.move( file_name, folder_name + "/" + new_file_name)



