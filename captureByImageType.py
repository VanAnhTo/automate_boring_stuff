import cv2
import openpyxl
import datetime

cam = cv2.VideoCapture(0)
wb = openpyxl.load_workbook('Capture.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')

cv2.namedWindow("test")

img_counter = 0
path = 'screen_capture'
sheet.cell('C3').value = '2'
i = 2


while True:
    ret, frame = cam.read()
    cv2.imshow("test", frame)
    if not ret:
        break
    k = cv2.waitKey(1)

    if k%256 == 27:
        # ESC pressed
        print("Escape hit, closing...")
        break
    
    elif k%256 == 32:
        # SPACE pressed
        
        d = sheet.cell('C3').value
        print(d)
        
        row = sheet.cell(row= int(d), column=2).value  
        print(row)

        now = datetime.datetime.now()

<<<<<<< HEAD:2.py
        img_name = now.strftime(path +"/" + "%Y%m%d_%H%M%S_" +row+".png")
=======
        img_name = now.strftime(path +"/" + "%Y%m%d_%H%M%S-" +row+".png")
>>>>>>> origin/master:captureByImageType.py

        cv2.imwrite(img_name, frame)   
        print("{} written!".format(img_name))
        sheet.cell('C3').value = str(int(d) + 1)

       
cam.release()

cv2.destroyAllWindows()
