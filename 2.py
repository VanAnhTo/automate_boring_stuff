import cv2
import openpyxl

cam = cv2.VideoCapture(0)
wb = openpyxl.load_workbook('Capture.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')

cv2.namedWindow("test")

img_counter = 0
path = 'screen_capture'
sheet.cell('C3').value = '2'
i = 2
popup = 'popup'
line = 'line'
divide = 'divide'


while True:
    ret, frame = cam.read()
    cv2.imshow("test", frame)
    if not ret:
        break
    k = cv2.waitKey(1)

    if k%256 == 27:
        # ESC pressed
        print("Escape hit, closing...")
        break
    
    elif k%256 == 32:
        # SPACE pressed
        
        d = sheet.cell('C3').value
        print(d)
        
        row = sheet.cell(row= int(d), column=2).value  
        print(row)

        if(row == 1):
            
            img_name = path +"/" + "opencv_frame_{}_popup.png".format(d)
            cv2.imwrite(img_name, frame)
                    
            print("{} written!".format(img_name))
                    
            sheet.cell('C3').value = str(int(d) + 1)
            a = sheet.cell('C3').value
            print (a)
            
        if (row == 2):
            
            img_name = path +"/" + "opencv_frame_{}_divide.png".format(d)
            cv2.imwrite(img_name, frame)
                    
            print("{} written!".format(img_name))
                    
            sheet.cell('C3').value = str(int(d) + 1)
            a = sheet.cell('C3').value
            print (a)

        if (row == 3):

            img_name = path +"/" + "opencv_frame_{}_line.png".format(d)
            cv2.imwrite(img_name, frame)
                    
            print("{} written!".format(img_name))
                    
            sheet.cell('C3').value = str(int(d) + 1)
            a = sheet.cell('C3').value
            print (a)
            
       
cam.release()

cv2.destroyAllWindows()
