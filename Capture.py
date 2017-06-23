import cv2
import openpyxl

cam = cv2.VideoCapture(0)
wb = openpyxl.load_workbook('Capture.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')

cv2.namedWindow("test")

img_counter = 0
path = 'screen_capture'

while True:
    ret, frame = cam.read()
    cv2.imshow("test", frame)
    if not ret:
        break
    k = cv2.waitKey(1)

    if k%256 == 27:
        # ESC pressed
        print("Escape hit, closing...")
        break
    
    elif k%256 == 32:
        # SPACE pressed
        i = 2
        count = sheet.cell('C3').value
        row = sheet.cell(row=i, column=2).value
        if ( row == 1):
            img_name = path +"/" + "popup_opencv_frame_{}.png".format(img_counter)
            cv2.imwrite(img_name, frame)
            print("{} written!".format(img_name))
            img_counter += 1
            count = count +1
            
        if (row == 2):
            img_name = path +"/" + "line_opencv_frame_{}.png".format(img_counter)
            cv2.imwrite(img_name, frame)
            print("{} written!".format(img_name))
            img_counter += 1
            count = count +1
            
        if (row == 3):
            img_name = path +"/" + "haft_opencv_frame_{}.png".format(img_counter)
            cv2.imwrite(img_name, frame)
            print("{} written!".format(img_name))
            img_counter += 1
            count = count +1
            
        i= i+1
        
cam.release()

cv2.destroyAllWindows()
