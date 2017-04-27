from ocr import ocr_space_file
import json
import openpyxl
import os
import shutil
import datetime

path = "input"
dirs = os.listdir( path )
    
wb = openpyxl.load_workbook('example.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')
i = 2
for file in dirs:
    file_name = file
    test_file = ocr_space_file(filename = path + "/" + file_name, language='jpn')
    json_obj = json.loads(test_file)
    print(json_obj)

    actual_result = ""        
    actual_result = json_obj["ParsedResults"][0]["ParsedText"].strip().replace("\r\n", "")
    print ('Actual: ' + actual_result)
    
    expected_result = sheet.cell(row=i, column=1).value
    
    
    sheet['B'+ str(i)] = actual_result
    print('Expected: ' + expected_result)
  
    if  expected_result == actual_result:
       sheet.cell(row=i, column=3).value = 'Pass'
    else:
       sheet.cell(row=i, column=3).value = 'Fail'
    print (sheet.cell(row=i, column=3).value)

    folder_name = "backup"
    os.makedirs(folder_name, exist_ok=True)
    now = datetime.datetime.now()
    new_file_name = now.strftime("%Y%m%d%H%M%S%f.png")
    print(new_file_name)
    shutil.move( path + "/" + file_name, folder_name + "/" + new_file_name)

    sheet.cell(row=i, column=4).hyperlink = folder_name + "/" + new_file_name
    

    i= i+1
    
wb.save('example.xlsx')


